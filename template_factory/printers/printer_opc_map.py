from __future__ import absolute_import

""" Template Factory: OPC-map XLS printer """


__author__     = "Krisztian Loki"
__copyright__  = "Copyright 2019, European Spallation Source, Lund"
__license__    = "GPLv3"



import os.path

from . import PRINTER
from tf_ifdef import BASE_TYPE



try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    def printer():
        return (OPC_MAP_XLS.name(), OPC_MAP_XLS)
except ImportError:
    def printer():
        raise NotImplementedError(OPC_MAP_XLS.name())



class OPC_MAP_XLS(PRINTER):
    TAG_NAME   = "A1"
    TAG_TYPE   = "B1"
    EPICS_NAME = "C1"
    EPICS_TYPE = "D1"

    def __init__(self):
        super(OPC_MAP_XLS, self).__init__(comments = False, preserve_empty_lines = False, show_origin = False)
        self._wb       = None
        self._filename = "opc-map.xlsx"


    @staticmethod
    def name():
        return "OPC-MAP.XLS"


    #
    # HEADER
    #
    def header(self, header_if_def, output, **keyword_params):
        super(OPC_MAP_XLS, self).header(header_if_def, output, **keyword_params)
        self._wb = Workbook()
        try:
            self._filename = self._helpers.sanitizeFilename(keyword_params["PLCF"].process(self.filename(extension = "xlsx")))
        except KeyError:
            pass


    def _initialize_ws(self, ws):
        ws.column_dimensions['A'].width = 90
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 90
        ws.column_dimensions['D'].width = 12
        ws[OPC_MAP_XLS.TAG_NAME]   = "Tag name"
        ws[OPC_MAP_XLS.TAG_TYPE]   = "Tag type"
        ws[OPC_MAP_XLS.EPICS_NAME] = "EPICS name"
        ws[OPC_MAP_XLS.EPICS_TYPE] = "EPICS type"
        ws[OPC_MAP_XLS.TAG_NAME].font   = Font(bold=True)
        ws[OPC_MAP_XLS.TAG_TYPE].font   = Font(bold=True)
        ws[OPC_MAP_XLS.EPICS_NAME].font = Font(bold=True)
        ws[OPC_MAP_XLS.EPICS_TYPE].font = Font(bold=True)


    #
    # BODY
    #
    def _ifdef_body(self, if_def, output, **keyword_params):
        for var in if_def.interfaces():
            if not isinstance(var, BASE_TYPE):
                continue

            datablock_name = self.expand(var.datablock_name()).replace(':', '_')
            try:
                ws = self._wb[datablock_name]
            except KeyError:
                ws = self._wb.create_sheet(datablock_name)
                self._initialize_ws(ws)

            cols = [ var.name(), var.plc_type(), self.create_pv_name(self.inst_slot(if_def), var), var.pv_type() ]
            try:
                cplcf = keyword_params["PLCF"]
                ws.append([ cplcf.process(x) for x in cols ])
            except KeyError:
                ws.append(cols)


    #
    # FOOTER
    #
    def footer(self, footer_if_def, output, **keyword_params):
        super(OPC_MAP_XLS, self).footer(footer_if_def, output, **keyword_params)
        if self._wb is None:
            return
        self._wb.remove(self._wb.active)
        self._wb.save(os.path.join(self._output_dir, self._filename))
